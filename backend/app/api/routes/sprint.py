from datetime import datetime, timedelta, date as date_type
import io

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse, Response
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import update, delete

from app.models.news import News
from app.api.routes.auth import get_current_user
from app.db.session import get_db
from app.models.project_member import ProjectMember
from app.models.sprint import Sprint
from app.models.sprint_status import SprintStatus
from app.models.sprint_task_event import SprintTaskEvent
from app.models.sprint_task_event_type import SprintTaskEventType
from app.models.task import Task
from app.models.task_workflow_status import TaskWorkflowStatus
from app.models.team import Team
from app.models.team_member import TeamMember
from app.models.user import User
from app.models.task_multiple_assignees import task_multiple_assignees
from app.services.news import create_news_for_users
from app.schemas.sprint import BurndownPoint, SprintCreate, SprintRead, SprintStatsRead
from app.services.sprint_burndown import (
    build_burndown_points,
    build_fallback_events,
    snapshot_story_points,
    summarize_sprint_tasks,
)
from app.services.sprint_status import sync_team_sprint_statuses
from app.services.task_delete import get_task_delete_state
from app.schemas.task import TaskRead
from fastapi.responses import JSONResponse
from fastapi.encoders import jsonable_encoder

router = APIRouter(prefix="/sprints", tags=["sprints"])


@router.get("", response_model=list[SprintRead])
async def get_sprints(
    team_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    team = await get_team_or_404(team_id, db)
    await get_project_membership_or_404(team.fk_projectid_project, current_user.id_user, db)
    await sync_team_sprint_statuses(team_id, db)

    result = await db.execute(select(Sprint).where(Sprint.fk_teamid_team == team_id))
    sprints = result.scalars().all()

    output = []
    for sprint in sprints:
        result = await db.execute(
            select(Task)
            .where(Task.fk_sprintid_sprint == sprint.id_sprint)
            .order_by(Task.board_order.asc(), Task.id_task.asc())
        )

        tasks = result.scalars().all()
        for t in tasks:
            await db.refresh(t)

        task_reads = []
        for task in tasks:
            # Load multi‑assignees
            assignees_result = await db.execute(
                select(task_multiple_assignees.fk_team_memberid_team_member)
                .where(task_multiple_assignees.fk_taskid_task == task.id_task)
            )
            assignees = [row[0] for row in assignees_result.fetchall()]

            # Build TaskRead
            can_delete, delete_block_reason = await get_task_delete_state(task, db)
            task_read = TaskRead.model_validate(task)
            task_read.can_delete = can_delete
            task_read.delete_block_reason = delete_block_reason
            task_read.assignees = assignees

            task_reads.append(task_read.model_dump())


        output.append(
            {
                "id_sprint": sprint.id_sprint,
                "start_date": sprint.start_date,
                "end_date": sprint.end_date,
                "status": sprint.status,
                "tasks": task_reads,
                "name": sprint.name,
            }
        )

    # ⭐ Convert everything to JSON‑safe types
    safe_output = jsonable_encoder(output)

    # ⭐ Disable caching
    return JSONResponse(
        content=safe_output,
        headers={"Cache-Control": "no-store"}
    )


@router.post("", response_model=dict)
async def create_sprint(
    payload: SprintCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    team = await get_team_or_404(payload.team_id, db)
    await get_project_membership_or_404(team.fk_projectid_project, current_user.id_user, db)
    today = datetime.utcnow().date()

    if payload.end_date < payload.start_date:
        raise HTTPException(
            status_code=400,
            detail="Sprint end date must be on or after the start date",
        )

    if payload.end_date.date() < today:
        raise HTTPException(
            status_code=400,
            detail="Sprint end date cannot be in the past",
        )

    # Count all sprints for this team to generate the name
    count_result = await db.execute(
        select(Sprint).where(Sprint.fk_teamid_team == payload.team_id)
    )
    sprint_count = len(count_result.scalars().all())
    sprint_name = f"Sprint {sprint_count + 1}"

    sprint = Sprint(
        fk_teamid_team=payload.team_id,
        start_date=payload.start_date,
        end_date=payload.end_date,
        status=SprintStatus.PLANNED.value,
        name=sprint_name,
    )

    db.add(sprint)
    await db.commit()
    await db.refresh(sprint)

    return {"status": "ok", "id_sprint": sprint.id_sprint}


@router.patch("/{sprint_id}")
async def update_sprint(
    sprint_id: int,
    payload: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    sprint = await db.get(Sprint, sprint_id)
    if not sprint:
        raise HTTPException(404, "Sprint not found")

    start_date_str = payload.get("start_date")
    end_date_str = payload.get("end_date")

    start_date = datetime.fromisoformat(start_date_str) if start_date_str else None
    end_date = datetime.fromisoformat(end_date_str) if end_date_str else None

    if start_date and end_date and start_date > end_date:
        raise HTTPException(400, "Start date cannot be after end date")

    # Apply updates
    if start_date:
        sprint.start_date = start_date
    if end_date:
        sprint.end_date = end_date
    if "name" in payload:
        sprint.name = payload["name"]

    # Only validate overlap if dates were changed
    effective_start = start_date or sprint.start_date
    effective_end = end_date or sprint.end_date

    if start_date or end_date:
        result = await db.execute(
            select(Sprint)
            .where(
                Sprint.fk_teamid_team == sprint.fk_teamid_team,
                Sprint.id_sprint != sprint.id_sprint,
                Sprint.status != SprintStatus.COMPLETED.value,
                effective_start <= Sprint.end_date,
                effective_end >= Sprint.start_date,
            )
        )
        overlapping = result.scalars().first()
        if overlapping:
            raise HTTPException(
                status_code=400,
                detail=f"Sprint dates overlap with Sprint {overlapping.id_sprint}",
            )

    db.add(sprint)
    await db.commit()
    await db.refresh(sprint)

    return {"message": "Sprint updated"}




@router.get("/{sprint_id}/stats", response_model=SprintStatsRead)
async def get_sprint_stats(
    sprint_id: int,
    team_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    team = await get_team_or_404(team_id, db)
    await get_project_membership_or_404(team.fk_projectid_project, current_user.id_user, db)
    await sync_team_sprint_statuses(team_id, db)

    result = await db.execute(
        select(Sprint).where(
            Sprint.id_sprint == sprint_id,
            Sprint.fk_teamid_team == team_id,
        )
    )
    sprint = result.scalar_one_or_none()

    if sprint is None:
        raise HTTPException(status_code=404, detail="Sprint not found")

    result = await db.execute(
        select(Task).where(Task.fk_sprintid_sprint == sprint.id_sprint)
    )
    tasks = result.scalars().all()

    sprint_length_days = max((sprint.end_date.date() - sprint.start_date.date()).days + 1, 1)

    if sprint.start_date.tzinfo is not None:
        today = datetime.now(sprint.start_date.tzinfo).date()
    else:
        today = datetime.utcnow().date()
    current_date = min(max(today, sprint.start_date.date()), sprint.end_date.date())
    elapsed_days = max((current_date - sprint.start_date.date()).days + 1, 0)
    remaining_days = max((sprint.end_date.date() - current_date).days, 0)
    event_result = await db.execute(
        select(SprintTaskEvent)
        .where(SprintTaskEvent.fk_sprintid_sprint == sprint.id_sprint)
        .order_by(SprintTaskEvent.occurred_at.asc(), SprintTaskEvent.id_sprint_task_event.asc())
    )
    task_events = event_result.scalars().all()
    if not task_events:
        task_events = build_fallback_events(tasks, sprint.start_date, sprint.end_date)

    burndown_points = build_burndown_points(sprint.start_date, sprint.end_date, task_events)
    summary = summarize_sprint_tasks(task_events, tasks)
    planned_points_per_day = (
        summary.committed_story_points / sprint_length_days if sprint_length_days else 0.0
    )
    completion_rate = (
        round((summary.completed_story_points / summary.committed_story_points) * 100, 1)
        if summary.committed_story_points > 0
        else 0.0
    )

    return SprintStatsRead(
        id_sprint=sprint.id_sprint,
        team_id=team_id,
        start_date=sprint.start_date,
        end_date=sprint.end_date,
        status=sprint.status,
        committed_tasks=summary.committed_tasks,
        committed_story_points=summary.committed_story_points,
        completed_tasks=summary.completed_tasks,
        completed_story_points=summary.completed_story_points,
        remaining_tasks=summary.remaining_tasks,
        remaining_story_points=summary.remaining_story_points,
        rolled_over_tasks=summary.rolled_over_tasks,
        rolled_over_story_points=summary.rolled_over_story_points,
        sprint_length_days=sprint_length_days,
        elapsed_days=min(elapsed_days, sprint_length_days),
        remaining_days=remaining_days,
        planned_points_per_day=round(planned_points_per_day, 2),
        completion_rate=completion_rate,
        burndown_points=burndown_points,
    )


@router.get("/{sprint_id}/export")
async def export_sprint(
    sprint_id: int,
    team_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    team = await get_team_or_404(team_id, db)
    await get_project_membership_or_404(team.fk_projectid_project, current_user.id_user, db)
    await sync_team_sprint_statuses(team_id, db)

    result = await db.execute(
        select(Sprint).where(
            Sprint.id_sprint == sprint_id,
            Sprint.fk_teamid_team == team_id,
        )
    )
    sprint = result.scalar_one_or_none()

    if sprint is None:
        raise HTTPException(status_code=404, detail="Sprint not found")

    if sprint.status != SprintStatus.COMPLETED.value:
        raise HTTPException(
            status_code=400,
            detail="Sprint export is only available for completed sprints",
        )

    # Get sprint tasks
    task_result = await db.execute(
        select(Task).where(Task.fk_sprintid_sprint == sprint.id_sprint)
    )
    tasks = task_result.scalars().all()

    # Get burndown data
    event_result = await db.execute(
        select(SprintTaskEvent).where(SprintTaskEvent.fk_taskid_task.in_([t.id_task for t in tasks]))
    )
    events = event_result.scalars().all()
    
    if not events:
        events = build_fallback_events(tasks, sprint.start_date, sprint.end_date)
    
    burndown_points = build_burndown_points(sprint.start_date, sprint.end_date, events)

    # Debug: Check if we have data
    print(f"Debug: Found {len(tasks)} tasks, {len(events)} events, {len(burndown_points)} burndown points")
    print(f"Debug: Burndown points sample: {burndown_points[:2] if burndown_points else 'None'}")

    # Get project teams and their members
    project_result = await db.execute(
        select(Team).where(Team.fk_projectid_project == team.fk_projectid_project)
    )
    project_teams = project_result.scalars().all()

    # Get team members for all teams in the project
    team_ids = [t.id_team for t in project_teams]
    team_member_result = await db.execute(
        select(TeamMember, User, Team).where(
            TeamMember.fk_teamid_team.in_(team_ids)
        ).join(User, TeamMember.fk_userid_user == User.id_user).join(Team, TeamMember.fk_teamid_team == Team.id_team)
    )
    team_members_data = team_member_result.all()

    # Get all tasks in the project for team statistics
    project_task_result = await db.execute(
        select(Task).where(
            Task.fk_projectid_project == team.fk_projectid_project,
            Task.fk_sprintid_sprint == sprint.id_sprint
        )
    )
    project_tasks = project_task_result.scalars().all()

    # Get team members with their teams and users
    team_member_result = await db.execute(
        select(TeamMember, User, Team).where(
            TeamMember.fk_teamid_team.in_(team_ids)
        ).join(User, TeamMember.fk_userid_user == User.id_user).join(Team, TeamMember.fk_teamid_team == Team.id_team)
    )
    team_members_data = team_member_result.all()

    # Create lookup dictionaries
    team_member_lookup = {tm.id_team_member: (tm, user, team) for tm, user, team in team_members_data}
    team_lookup = {t.id_team: t for t in project_teams}

    try:
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Sprint Summary"

        # Sprint Summary Sheet
        summary_sheet.append(["Field", "Value"])
        summary_sheet.append(["Sprint ID", sprint.id_sprint])
        summary_sheet.append(["Team ID", sprint.fk_teamid_team])
        summary_sheet.append(["Start Date", sprint.start_date.isoformat()])
        summary_sheet.append(["End Date", sprint.end_date.isoformat()])
        summary_sheet.append(["Status", sprint.status])
        summary_sheet.append(["Total Tasks", len(tasks)])
        summary_sheet.append(["Completed Tasks", sum(1 for task in tasks if task.workflow_status == TaskWorkflowStatus.DONE.value)])
        summary_sheet.append(["Committed Story Points", sum((task.story_points or 0) for task in tasks)])
        summary_sheet.append(["Completed Story Points", sum((task.story_points or 0) for task in tasks if task.workflow_status == TaskWorkflowStatus.DONE.value)])

        # Team Statistics Sheet
        team_stats_sheet = workbook.create_sheet("Team Statistics")
        team_stats_sheet.append(["Team Name", "Total Story Points", "Completed Story Points", "Completion Rate (%)"])

        team_stats = {}
        for task in project_tasks:
            team_name = "Unassigned"
            if task.fk_teamid_team and task.fk_teamid_team in team_lookup:
                team_name = team_lookup[task.fk_teamid_team].name or f"Team {task.fk_teamid_team}"
            
            if team_name not in team_stats:
                team_stats[team_name] = {"total": 0, "completed": 0}
            
            story_points = task.story_points or 0
            team_stats[team_name]["total"] += story_points
            if task.workflow_status == TaskWorkflowStatus.DONE.value:
                team_stats[team_name]["completed"] += story_points

        for team_name, stats in team_stats.items():
            completion_rate = (stats["completed"] / stats["total"] * 100) if stats["total"] > 0 else 0
            team_stats_sheet.append([
                team_name,
                stats["total"],
                stats["completed"],
                round(completion_rate, 2)
            ])

        # Team Member Statistics Sheet
        member_stats_sheet = workbook.create_sheet("Team Member Statistics")
        member_stats_sheet.append(["Team Name", "Member Name", "Email", "Assigned Story Points", "Completed Story Points", "Completion Rate (%)"])

        member_stats = {}
        for task in project_tasks:
            if not task.fk_team_memberid_team_member or task.fk_team_memberid_team_member not in team_member_lookup:
                continue
            
            team_member, user, team = team_member_lookup[task.fk_team_memberid_team_member]
            team_name = team.name or f"Team {team.id_team}"
            member_key = f"{team_name}_{user.id_user}"
            
            if member_key not in member_stats:
                member_stats[member_key] = {
                    "team_name": team_name,
                    "name": user.name or "Unknown",
                    "email": user.email,
                    "total": 0,
                    "completed": 0
                }
            
            story_points = task.story_points or 0
            member_stats[member_key]["total"] += story_points
            if task.workflow_status == TaskWorkflowStatus.DONE.value:
                member_stats[member_key]["completed"] += story_points

        for stats in member_stats.values():
            completion_rate = (stats["completed"] / stats["total"] * 100) if stats["total"] > 0 else 0
            member_stats_sheet.append([
                stats["team_name"],
                stats["name"],
                stats["email"],
                stats["total"],
                stats["completed"],
                round(completion_rate, 2)
            ])

        # Sprint Tasks Sheet
        task_sheet = workbook.create_sheet("Sprint Tasks")
        task_sheet.append([
            "Task ID",
            "Name",
            "Status",
            "Story Points",
            "Completed At",
            "Board Order",
            "Description",
            "Assigned Team",
            "Assigned Member",
        ])

        for task in tasks:
            # Find team and member info for this task
            team_name = "Unassigned"
            member_name = "Unassigned"
            
            if task.fk_teamid_team and task.fk_teamid_team in team_lookup:
                team_name = team_lookup[task.fk_teamid_team].name or f"Team {task.fk_teamid_team}"
            
            if task.fk_team_memberid_team_member and task.fk_team_memberid_team_member in team_member_lookup:
                team_member, user, _ = team_member_lookup[task.fk_team_memberid_team_member]
                member_name = user.name or "Unknown"
            
            task_sheet.append([
                task.id_task,
                task.name or "",
                task.workflow_status,
                task.story_points or 0,
                task.completed_at.isoformat() if task.completed_at else "",
                task.board_order,
                task.description or "",
                team_name,
                member_name,
            ])

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        print(f"Debug: Excel file created successfully, size: {len(buffer.getvalue())} bytes")
        
        excel_data = buffer.getvalue()

    except Exception as e:
        print(f"Debug: Error creating Excel file: {e}")
        raise HTTPException(status_code=500, detail=f"Error creating Excel file: {str(e)}")

    filename = f"sprint-{sprint.id_sprint}-report.xlsx"
    return Response(
        content=excel_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
        },
    )


@router.post("/{sprint_id}/start")
async def start_sprint(
    sprint_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(Sprint).where(Sprint.id_sprint == sprint_id))
    sprint = result.scalar_one_or_none()

    if sprint is None:
        raise HTTPException(status_code=404, detail="Sprint not found")

    team = await get_team_or_404(sprint.fk_teamid_team, db)
    await get_project_membership_or_404(team.fk_projectid_project, current_user.id_user, db)

    if sprint.status != SprintStatus.PLANNED.value:
        raise HTTPException(status_code=400, detail="Only a PLANNED sprint can be started")

    sprint.status = SprintStatus.ACTIVE.value
    db.add(sprint)

    member_result = await db.execute(
        select(TeamMember.fk_userid_user)
        .where(TeamMember.fk_teamid_team == team.id_team)
    )
    member_ids = member_result.scalars().all()
    if member_ids:
        await create_news_for_users(
            db=db,
            user_ids=member_ids,
            title="Sprint started",
            message=f"Sprint {sprint.id_sprint} has started for team {team.name}.",
            news_type="sprint_started",
            project_id=team.fk_projectid_project,
            team_id=team.id_team,
            sprint_id=sprint.id_sprint,
        )

    await db.commit()
    await db.refresh(sprint)

    return {"status": "ok", "id_sprint": sprint.id_sprint}


@router.post("/{sprint_id}/close")
async def close_sprint(
    sprint_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(Sprint).where(Sprint.id_sprint == sprint_id))
    sprint = result.scalar_one_or_none()

    if sprint is None:
        raise HTTPException(status_code=404, detail="Sprint not found")

    team = await get_team_or_404(sprint.fk_teamid_team, db)
    await get_project_membership_or_404(team.fk_projectid_project, current_user.id_user, db)

    if sprint.status == SprintStatus.COMPLETED.value:
        raise HTTPException(status_code=400, detail="Sprint is already completed")

    unfinished_result = await db.execute(
        select(Task)
        .where(
            Task.fk_sprintid_sprint == sprint.id_sprint,
            Task.workflow_status != TaskWorkflowStatus.DONE.value,
        )
        .order_by(Task.id_task.asc())
    )
    unfinished_tasks = unfinished_result.scalars().all()

    next_sprint_result = await db.execute(
        select(Sprint)
        .where(
            Sprint.fk_teamid_team == sprint.fk_teamid_team,
            Sprint.id_sprint != sprint.id_sprint,
            Sprint.start_date > sprint.end_date,
            Sprint.status != SprintStatus.COMPLETED.value,
        )
        .order_by(Sprint.start_date.asc(), Sprint.id_sprint.asc())
    )




    next_sprint = next_sprint_result.scalars().first()

    if next_sprint is None:
        count_result = await db.execute(
            select(Sprint).where(Sprint.fk_teamid_team == sprint.fk_teamid_team)
        )
        sprint_count = len(count_result.scalars().all())
        sprint_name = f"Sprint {sprint_count + 1}"

        duration = sprint.end_date - sprint.start_date
        new_start = sprint.end_date + timedelta(days=1)
        new_end = new_start + duration
        next_sprint = Sprint(
            fk_teamid_team=sprint.fk_teamid_team,
            start_date=new_start,
            end_date=new_end,
            status=SprintStatus.PLANNED.value,
            name=sprint_name,
        )
        db.add(next_sprint)
        await db.flush()

    next_board_order = 0
    if next_sprint is not None:
        next_order_result = await db.execute(
            select(Task)
            .where(
                Task.fk_sprintid_sprint == next_sprint.id_sprint,
                Task.workflow_status == TaskWorkflowStatus.TODO.value,
            )
            .order_by(Task.board_order.desc(), Task.id_task.desc())
        )
        existing_todo_tasks = next_order_result.scalars().all()
        if existing_todo_tasks:
            next_board_order = (existing_todo_tasks[0].board_order or 0) + 1

    moved_task_ids: list[int] = []
    occurred_at = datetime.utcnow()

    for task in unfinished_tasks:
        add_sprint_event = SprintTaskEvent(
            fk_taskid_task=task.id_task,
            fk_sprintid_sprint=sprint.id_sprint,
            event_type=SprintTaskEventType.REMOVED.value,
            story_points=snapshot_story_points(task),
            occurred_at=occurred_at,
        )
        db.add(add_sprint_event)

        task.fk_sprintid_sprint = next_sprint.id_sprint if next_sprint is not None else None
        task.workflow_status = TaskWorkflowStatus.TODO.value
        task.board_order = next_board_order if next_sprint is not None else 0
        if next_sprint is not None:
            next_board_order += 1
            db.add(
                SprintTaskEvent(
                    fk_taskid_task=task.id_task,
                    fk_sprintid_sprint=next_sprint.id_sprint,
                    event_type=SprintTaskEventType.ADDED.value,
                    story_points=snapshot_story_points(task),
                    occurred_at=occurred_at,
                )
            )

        db.add(task)
        moved_task_ids.append(task.id_task)

    sprint.status = SprintStatus.COMPLETED.value
    db.add(sprint)

    await db.commit()
    await db.refresh(sprint)

    return {
        "status": "ok",
        "sprint_id": sprint.id_sprint,
        "sprint_status": sprint.status,
        "moved_task_ids": moved_task_ids,
        "target_sprint_id": next_sprint.id_sprint if next_sprint is not None else None,
    }


async def get_team_or_404(team_id: int, db: AsyncSession):
    result = await db.execute(select(Team).where(Team.id_team == team_id))
    team = result.scalar_one_or_none()

    if team is None:
        raise HTTPException(status_code=404, detail="Team not found")

    return team


async def get_project_membership_or_404(project_id: int, user_id: int, db: AsyncSession):
    result = await db.execute(
        select(ProjectMember).where(
            ProjectMember.fk_projectid_project == project_id,
            ProjectMember.fk_userid_user == user_id,
        )
    )
    membership = result.scalar_one_or_none()

    if membership is None:
        raise HTTPException(status_code=404, detail="Project not found")

    return membership


@router.delete("/{sprint_id}")
async def delete_sprint(
    sprint_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    sprint = await db.get(Sprint, sprint_id)
    if not sprint:
        raise HTTPException(404, "Sprint not found")

    # Load team and check membership
    team = await get_team_or_404(sprint.fk_teamid_team, db)
    await get_project_membership_or_404(
        team.fk_projectid_project,
        current_user.id_user,
        db
    )

    #  Delete sprint events
    await db.execute(
        delete(SprintTaskEvent)
        .where(SprintTaskEvent.fk_sprintid_sprint == sprint_id)
    )

    #  Delete news entries referencing this sprint
    await db.execute(
        delete(News)
        .where(News.fk_sprintid_sprint == sprint_id)
    )

    #  Move tasks to backlog
    await db.execute(
        update(Task)
        .where(Task.fk_sprintid_sprint == sprint_id)
        .values(fk_sprintid_sprint=None)
    )

    #  Delete sprint
    await db.delete(sprint)
    await db.commit()

    return {"message": "Sprint deleted"}